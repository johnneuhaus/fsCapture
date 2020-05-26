#! /usr/bin/env python3
# -*- coding: utf-8 -*-
# This script is used to control a Rohde & Schwarz ETL
# Spectrum Analyzer.

# This script is:
# C:\Shared\batch_files\fsCapture.py

# Author: John Neuhaus, jneuhausATosborn-engDOTcom
# 2020-05-22 11:43

# Requires additional packages to be installed.
# Issue the following at a CMD prompt:
#   pip install openpyxl
#   pip install pyvisa
#   pip install pyvisa-py
#   pip install ntplib


'''
MIT License

Copyright (c) 2020 John Neuhaus, jneuhausATosborn-engDOTcom

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
'''

import time
import datetime
import re
import os
import sys
import configparser

FMT = '%Y-%m-%d %X ' # Time stamp format for CONSOLE messages

# ———————————————————————————————————————————————————
#               Set Variables
# ———————————————————————————————————————————————————

def setvars():
    
    global _entryError
    _entryError = ''  # Message displayed when user made invalid choice

    global INI_FILENAME    
    global inibasename
    INI_FILENAME = os.path.join(LOG_PATH, inibasename)
    if len(sys.argv) > 2 : 
        # Specify alternate INI file as second arg on command line
        logger.debug('Second argument on command line is ' + sys.argv[2] )
        ini_file_spec = os.path.join(LOG_PATH, inibasename)
        if not os.path.exists(ini_file_spec) :  
            # Specified file does not exist
            # Menu for user to decide what to do if specified file does 
            #   not exist
            choice ='0'
            while choice =='0':
                inputPrompt = ("""              File name specified was: """ \
                    + inibasename + """
              File does not exist in """ + LOG_PATH + """

              1 — Create this file \'""" + inibasename + """\'
              2 — Enter a different file name
              Q — Quit
              """ + _entryError + """
              Type a choice and press ENTER""")
                choice = input (inputPrompt + ' >> ')
                _entryError = ''
                if choice == 'q':
                    done()
                elif choice == 'Q':
                    done()
                elif choice == '1':
                    INI_FILENAME = ini_file_spec
                elif choice == '2':
                    inibasename = input ('Enter new file name >> ')
                    ini_file_spec = os.path.join(LOG_PATH, inibasename)
                    if not os.path.exists(ini_file_spec):
                        choice = '0' 
                    else:
                        INI_FILENAME = ini_file_spec
                #elif choice == '3':
                #    pass
                else:
                    _entryError = \
                        'Your entry "'+choice+'" is not a valid choice'
                    choice = '0' 
        else:
            INI_FILENAME = ini_file_spec  # Specified file exists so use it
    logger.warning('INI file is ' + INI_FILENAME )
    # Following are user-entered and stored in INI file:
    global ipaddr
    ipaddr = ''
    global _callsign
    _callsign = "XXXX"
    global _channel
    _channel = "8"
    global _radial
    _radial = "xxx"
    global _dist
    _dist = "xx"
    global _testposition
    _testposition = "B"
    global _truckheading
    _truckheading = ""
    global _temperature
    _temperature = ""
    global _wind 
    _wind = ""
    global _skycond
    _skycond = ""
    global _precip
    _precip = ""
    global _techname
    _techname = ""
    global _clutter
    _clutter = ""
    global _antdirup
    _antdirup = ""
    global _antdirdown
    _antdirdown = ""
    global _checklist_template
    _checklist_template = \
"C:\Shared\Info and Template\Folder Template\Site Info\Template Checklist R2.xlsx"
    global _default_meas_results_folder
    _default_meas_results_folder = 'Z:\\Measurement_results'
    global _meas_results_folder
    _meas_results_folder = _default_meas_results_folder
    global _default_channel_table
    _default_channel_table = "TV-USA-ATSC"
    global _channel_table
    _channel_table = _default_channel_table
    # End of user-entered variables.

    if not os.path.exists(INI_FILENAME):
        logger.debug('INI file ' + INI_FILENAME + ' does not exist.')
    else:
        logger.debug('Reading from ' + \
            str(config.read(INI_FILENAME)).replace('\\\\','\\'))
        logger.debug('Sections in INI_FILENAME: '+ str(config.sections()))
        section = 'UserEntered'
        if section in config: # If section exists in INI file
            try:
                _callsign = config.get(section,'_callsign')
            except:
                pass
            try:
                _channel = config.get(section,'_channel')
            except:
                pass
            try:
                _radial = config.get(section,'_radial')
            except:
                pass
            try:
                _dist = config.get(section,'_dist')
            except:
                pass
            try:
                _testposition = config.get(section,'_testposition')
            except:
                pass
            try:
                _truckheading = config.get(section,'_truckheading')
            except:
                pass
            try:
                _temperature = config.get(section,'_temperature')
            except:
                pass
            try:
                _wind = config.get(section,'_wind')
            except:
                pass
            try:
                _skycond = config.get(section,'_skycond')
            except:
                pass
            try:
                _precip = config.get(section,'_precip')
            except:
                pass
            try:
                _techname = config.get(section,'_techname')
            except:
                pass
            try:
                _clutter = config.get(section,'_clutter')
            except:
                pass
            try:
                _antdirup = config.get(section,'_antdirup')
            except:
                pass
            try:
                _antdirdown = config.get(section,'_antdirdown')
            except:
                pass
            try:
                _channel_table = config.get(section, '_channel_table')
            except:
                pass
        try:
            _checklist_template = config.get('ChecklistTemplate',\
                '_checklist_template')
        except:
            logger.debug('Did not read _checklist_template from INI file')
            pass
        try:
            _meas_results_folder = config.get('Measurement_Results_Folder',\
                '_meas_results_folder')
        except:
            logger.debug('Did not read _meas_results_folder from INI file')
            pass

        logger.debug('After reading INI file, _callsign is ' + _callsign)
        logger.debug('After reading INI file, _testposition is ' + \
            _testposition )
        logger.debug('After reading INI file, _checklist_template is ' + \
            _checklist_template)
        logger.debug('After reading INI file, _meas_results_folder is ' + \
            _meas_results_folder)

    global _MEASPT
    _MEASPT = "R" + _radial + "M" + _dist
    
    global _RevRadial
    _RevRadial = ""
    global _AzToTx
    _AzToTx = ""
    global _file
    _file = 'SetByUser'
    global _ext
    _ext = 'PNG'
    global _lat
    _lat = ''
    global _lon
    _lon = ''
    global _altitude
    _altitude = ''
    global _frequency
    _frequency = 0
    global timedatewarning
    timedatewarning = ''



# ———————————————————————————————————————————————————
#              Logging
# ———————————————————————————————————————————————————

def request_path(inipath,inibasename):
    # Menu for user to decide what to do if specified file does not exist
    
    global LOG_PATH
    #global inibasename

    choice ='0'
    ini_file_spec = os.path.join(inipath,inibasename)
    while choice =='0':
        _entryError = ''
        inputPrompt = ("""        Path for INI file does not exist: 
        \'""" + inipath + """\'

        1 — Create the path \'""" + inipath + """\'
        2 — Enter a different path
        3 — Use default \'"""+ LOG_PATH +"""\'
        Q — Quit
        """ + _entryError + """
        Type a choice and press ENTER""")
        choice = input (inputPrompt + ' >> ')
        _entryError = ''
        if choice == 'q':
            sys.exit(0)
        elif choice == 'Q':
            sys.exit(0)
        elif choice == '1':
            LOG_PATH = inipath
            try:
                os.mkdir(LOG_PATH)
            except OSError:
                #print('     Failed to create '+LOG_PATH)
                choice = '0'
                pass
        elif choice == '2':
            inipath = os.path.expandvars(input ('Enter new path >> ') )
            if not os.path.exists(inipath):
                choice = '0' 
            else:
                LOG_PATH = inipath
        elif choice == '3':
            try:
                os.mkdir(LOG_PATH)
            except OSError:
                #print('     Failed to create '+LOG_PATH)
                choice = '0'
                pass
        else:
            _entryError = 'Your entry "'+choice+'" is not a valid choice'
            choice = '0'
    return LOG_PATH




def start_logging():
    global LOG_FILENAME , LOG_PATH , inibasename
    #global LOG_PATH 
    #global inibasename
    LOG_PATH = os.path.expandvars(r'%APPDATA%')
    LOG_PATH = os.path.join(LOG_PATH,'Osborn')
    inipath = LOG_PATH
    inibasename = 'fsCapture.INI'

    if len(sys.argv) > 2 :  
        # Specify alternate INI file as second arg on command line
        inipath, inibasename = os.path.split(sys.argv[2])

    try:
        fileexists = os.path.exists(inipath) # Not exists = 0 or False
        LOG_PATH = inipath
    except FileNotFoundError:
        fileexists = False
        pass

    if fileexists == False:
        LOG_PATH = request_path(inipath,inibasename)

    # Specified path exists so use it
 
    LOG_FILENAME = os.path.join(LOG_PATH , 'fsCapture.log')

    import logging 
    from logging import handlers
    
    LEVELS = {'debug': logging.DEBUG,
            'info': logging.INFO,
            'warning': logging.WARNING,
            'error': logging.ERROR,
            'critical': logging.CRITICAL}

    if len(sys.argv) > 1:
        level_name = sys.argv[1]
        level = LEVELS.get(level_name, logging.NOTSET)
    else:
        level = logging.INFO
    # create logger
    global logger
    logger = logging.getLogger("fsCapture")
    logger.setLevel(level)


    # create file handler 
    #fh = logging.FileHandler(LOG_FILENAME)
    fh = logging.handlers.RotatingFileHandler\
        (LOG_FILENAME,maxBytes=50000,backupCount=5)
    fh.setLevel(level)
    # create console handler and set level to WARNING
    ch = logging.StreamHandler()
    ch.setLevel(logging.WARNING)
    # create formatter and add it to the handlers
    formatter = logging.Formatter\
        ("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
    consoleformatter = logging.Formatter("%(asctime)s - %(message)s", FMT)
    ch.setFormatter(consoleformatter)
    fh.setFormatter(formatter)
    # add the handlers to logger
    logger.addHandler(ch)
    logger.addHandler(fh)

        
    modificationTime = datetime.datetime.fromtimestamp\
        (os.path.getmtime(__file__)).strftime(FMT)
    logger.info('*** ' + __file__ + ' Mod: ' + modificationTime +' ***')

    
    logger.debug('logger is ' + str(logger))


# ———————————————————————————————————————————————————
#               INI (or config) file
# ———————————————————————————————————————————————————
config = configparser.ConfigParser()

def recall_setting(INI_FILENAME,section,keyword):
    if not os.path.exists(INI_FILENAME):
        logger.info(INI_FILENAME + ' does not exist.')
        raise
    else:
        config.read(INI_FILENAME)
        # Check if INI file has section and keyword
        try:
            keyvalue = config.get(section, keyword)
            return keyvalue
        # If it doesn't i.e. An exception was raised
        except configparser.NoOptionError:
            logger.debug("Could not get " + keyword + " from " + \
                INI_FILENAME)
            raise

def update_config_file(INI_FILENAME, section, keyword, keyvalue):
    config.read(INI_FILENAME)
    try:
        config.get(section, keyword)
    except configparser.NoSectionError:
        config.add_section(section)
    except configparser.NoOptionError:
        pass
    cfgfile = open(INI_FILENAME, 'w')
    config.set(section, keyword, keyvalue)
    config.write(cfgfile)
    #cfgfile.close()

def request_ipaddr():
    # Prompt user to type IP address and store it in INI file
    global ipaddr
    inputPrompt = 'Enter the IP address of the ETL analyzer'
    ipaddr = input (inputPrompt + ' >> ')
    import ipaddress
    try:
        ip = ipaddress.ip_address(ipaddr)
        logger.warning('%s is a valid IP%s address.' % (ip, ip.version))
        section = 'DEFAULT'
        keyword = 'ipaddr'
        keyvalue = ipaddr
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
        return
    except ValueError:
        logger.warning('Address is invalid: %s' % ipaddr )
    except:
        logger.warning('An exception occurred when validating IP address')
        pass

def recall_ipaddr():
    global ipaddr
    section , keyword = 'DEFAULT' , 'ipaddr'
    try:
        ipaddr = recall_setting(INI_FILENAME,section,keyword)
    except:
        request_ipaddr()
    if not ipaddr == '' :
        logger.info('IP Address in INI file is ' + str(ipaddr) )
    else:
        logger.info('IP address in INI file is blank')
        request_ipaddr()
 
# ———————————————————————————————————————————————————————————————
#      To run outside R&S Forum
# ———————————————————————————————————————————————————————————————
# 

#Bring in the VISA library
import pyvisa as visa
#visa.log_to_screen()  # debug: show details of all operations

#Create a resource manager
rm = visa.ResourceManager('@py')  #select the PyVISA-py backend using @py 
#rm = visa.ResourceManager()    # use default NI-VISA backend

def ping(host):
    global DEVICE
    try:
        DEVICE = rm.open_resource('TCPIP::' + ipaddr + '::INSTR')
        logger.debug("Connect to ("+ipaddr+") — True")
        DEVICE.close()
        return True
    except:
        logger.debug('Could not connect to ' + str(DEVICE) )
        return False


# ———————————————————————————————————————————————————
#              Check computer time 
# ———————————————————————————————————————————————————
def checkcomputertime() :
    
    import ntplib
    timeserver = 'us.pool.ntp.org'
    counter = 0
    ntpresponse = ''
    tdelta = 0
    while counter < 2 :
        counter += 1
        try:    
            ntpresponse = ntplib.NTPClient().request(timeserver).tx_time
            # Transmit NTP timestamp in system time
            counter = 5 # Success! Exit loop.
        except :
            logger.warning('NTP attempt no. ' +str(counter) + ' failed' )
            pass
            
    comp_time = datetime.datetime.now().timestamp()

    try:
        tdelta =  ntpresponse - comp_time
        logger.warning('Difference between NTP time and computer time: '+\
            str(round(tdelta,1)) + ' sec.')

    except:
        logger.warning('Failed to check computer time against NTP, number of attempts: '+\
            str(counter))
        pass

    if abs(tdelta) > 1 :
        logger.warning('Difference between NTP time and computer time: '+ \
            str(round(tdelta,1)) + ' sec.')

# ———————————————————————————————————————————————————
#              Verify connection to DEVICE instrument
# ———————————————————————————————————————————————————
def testconnectivity():
    global _etlIDN
    global ipaddr
    global tdelta
    logger.debug("testconnectivity() calling recall_ipaddr()")
    recall_ipaddr()
    logger.debug("testconnectivity() IP address is " + ipaddr )
    if (ping(ipaddr)) :
        logger.debug("testconnectivity() — received response to ping")
        withoutConnection = 0
        DEVICE.open()
        DEVICE.clear()
        DEVICE.write_termination = '\n'
        DEVICE.read_termination = '\n'
        DEVICE.write('*CLS', ) # Clear the Error queue
        try:
            _etlIDN = DEVICE.query('*IDN?', )
        except:
            pass

        logger.info('Connected to ' + _etlIDN + ' at IP addr ' + ipaddr)
        
        try:
            ETLdate = DEVICE.query('SYST:DATE?',) # 2020,5,9
        except:
            pass
        try:
            ETLtime = DEVICE.query('SYST:TIME?',) # 7,38,53
        except:
            pass
        
        ETLdatetime = str(datetime.datetime.strptime(ETLdate + ' ' + \
            ETLtime,'%Y,%m,%d %H,%M,%S'))
        # ETLdatetime = str(datetime.datetime.strptime(ETLdate + ' ' + \
        #   ETLtime,'%Y,%m,%d\n %H,%M,%S\n')) 
        # Why doesn't this work here?
        tdelta = datetime.datetime.strptime(ETLdatetime,'%Y-%m-%d %X').\
            timestamp() - datetime.datetime.now().timestamp()
        logger.debug('tdelta is ' + str(tdelta))
        global timedatewarning
        timedatewarning = ''
        if (abs(tdelta) > 2) :
            timedatewarning = '''
   ———— WARNING ————    ETL time is ''' + str(round(tdelta,1)) + \
       ''' seconds different from computer. 
              N — Compare computer time to NTP server
              1 — Set ETL date/time from computer'''
        logger.info('ETL date and time: ' + ETLdatetime + '   ' + \
            timedatewarning ,)

        # Operation complete query. Returns an ASCII "+1" 
        # when all pending overlapped operations have been completed
        try:
            _null = '0'
            _null = DEVICE.query('*OPC?',)
        except:
            logger.debug('testconnectivity() failed')
            pass
        logger.debug\
            ('testconnectivity() Operation Complete (OPC?) response: ' + \
                _null )
        DEVICE.close()
        
    else:
        logger.warning("No response from " + ipaddr )
        raise
  
    return


# ———————————————————————————————————————————————————
#               IP OK?
# ———————————————————————————————————————————————————
def ip_ok() :
    global withoutConnection
       
    try:
        testconnectivity()    
    except:  # Failed to connect
        choice ='0'
        while choice =='0':
                 
            inputPrompt = """Failed to open connection to """ + \
                str(ipaddr).replace('\n','') + """.

            1 — continue without connection or
            2 — enter a new IP address.
            
            Type choice and ENTER or just ENTER to Quit"""
            logger.debug(inputPrompt)
            choice = input (inputPrompt + ' >> ')
            
            if choice == "1":
                logger.warning\
                    ('User chose 1 to continue without connection.')
                withoutConnection = 1
                menu()
            if choice == "2":
                logger.debug('User chose 2 to enter new IP address.')
                withoutConnection = 1
                request_ipaddr()
                ip_ok()
                menu()
            else:
                logger.debug('User chose to quit.')
                pass  # Keep going
                
            done()
        
    else: 
        withoutConnection = 0
        DEVICE.close()
        return
    return
        


def gps(): 
    
    global _gps
    _gps = 0
    global _entryError
    
    _triedCount = 0
    _tryToConnect = 0
   
    if (ping(ipaddr)) :
        # currently connected to the GPS receiver?
        DEVICE.open()
        # when Radio/TV analyzer is not selected on ETL,
        # following throws VI_ERROR_TMO.
        try:
            _gps = int(DEVICE.query('SYST:POS:GPS:CONN?',))
            
            logger.debug('_gps is ' + str(_gps) )
        
            if _gps == 1 :
                # Displays GPS data in the Overview measurement.
                DEVICE.write('DISP:MEAS:OVER:GPS:STAT ON',)
                # Latitude:
                global _lat
                _lat = DEVICE.query('SYST:POS:LAT?',)
                # Longitude:
                global _lon
                _lon = DEVICE.query('SYST:POS:LONG?',)
                # Altitude in meters:
                global _altitude
                _altitude = DEVICE.query('SYST:POS:ALT?',)
                # Query GPS number of satellites:
                global _gpsValid
                _gpsValid = DEVICE.query('SYSTem:POSition:GPS:SATellites?',)
                
                # Operation complete query. Returns an ASCII "+1" 
                # when all pending overlapped operations have been 
                # completed
                _null = DEVICE.query('*OPC?',)
                DEVICE.close()
                
                logger.debug('Lat, Long: '+_lat.replace('\n','') + \
                    ','+_lon.replace('\n','') +\
                    ', Altitude: '+_altitude.replace('\n','') +\
                    ', GPS: ' + _gpsValid.replace('\n','') + \
                    ' (no. of satellites)'      )

            else:
                logger.warning(' ———— GPS is NOT CONNECTED! ————')
                DEVICE.close()
        except :
            _entryError = \
                'Did not get GPS status. ETL may not be in \
"TV/Radio Analyzer" mode.'
            logger.info(_entryError)    
            DEVICE.close()
            pass
        
    else:

        _entryError = 'FAILED to connect to ' + ipaddr + ''
        logger.info(_entryError)    

    
    return


# ———————————————————————————————————————————————————
#             MEASUREMENT RESULTS FOLDER
# ———————————————————————————————————————————————————
def meas_results_folder() :
    global withoutConnection
    global _default_meas_results_folder
    global _meas_results_folder
    if withoutConnection == 0 :  # If not connected (1), skip this
        # Test if _meas_results_folder exists
        logger.debug('Test if measurement results folder exists: ' + \
            _meas_results_folder )
        new_meas_results_folder = _meas_results_folder
        choice = '0'
        _entryError = ''
        while choice =='0':
            try:
                DEVICE.open()
                dir_list = DEVICE.query("MMEM:CAT? \'"+ \
                    new_meas_results_folder +"\'")
                DEVICE.close()
                choice = '1'
                _meas_results_folder = new_meas_results_folder
                logger.debug\
                    ('ETL reported that Measurement Results Folder \'' + \
                        new_meas_results_folder + '\' exists')
                section , keyword = 'Measurement_Results_Folder' ,\
                     '_meas_results_folder'
                keyvalue = _meas_results_folder
                update_config_file(INI_FILENAME, section, keyword, keyvalue)
            except:                
                DEVICE.close()
                choice ='0'
                inputPrompt = \
                    """  ETL reported that Measurement Results Folder \'"""\
                         + new_meas_results_folder + """\' does NOT exist.
                
                1 — continue with Measurement Results Folder that does not \
exist
                2 — enter a new Measurement Results Folder
                3 — use default \'""" + _default_meas_results_folder + \
                    """\'
                Q — Quit
                """ + _entryError + """
                Type a choice and press ENTER or ENTER alone to try again"""
                logger.debug(inputPrompt)
                choice = input (inputPrompt + ' >> ')
                _entryError = ''

                if choice == "q":
                    logger.debug('User chose to quit.')
                    done()
                elif choice == "Q":
                    logger.debug('User chose to quit.')
                    done()
            
                elif choice == "1":
                    logger.warning('User chose 1 to continue with \
Measurement Results Folder that does not exist.')
                elif choice == "2":
                    logger.debug\
                        ('User chose 2 to enter a new Measurement Results \
Folder.')
                    inputPrompt = 'New folder'
                    new_meas_results_folder = input (inputPrompt + ' >> ')
                    # Delete characters that are not legal in path names
                    new_meas_results_folder = re.sub(r'[*?<>|]',"",\
                        new_meas_results_folder)
                    choice = '0'
                elif choice == "3":
                    new_meas_results_folder = _default_meas_results_folder
                    logger.debug('User chose 3 to use default \'' + \
                        new_meas_results_folder + '\'.')
                    choice = '0'
                elif choice == "":
                    new_meas_results_folder = _default_meas_results_folder
                    logger.debug('User chose to try again ')
                    choice = '0'
                else:
                    _entryError = 'Your entry "'+choice+\
                        '" is not a valid choice'
                    choice = '0' 
                pass  # Keep going


# ———————————————————————————————————————————————————
#             CHANNEL TABLE
# ———————————————————————————————————————————————————
def channel_table() :
    global withoutConnection
    global _default_channel_table
    global _channel_table
    
    if withoutConnection == 0 :  # If not connected (1), skip this
        # Test if _channel_table exists

        new_channel_table = _channel_table
        choice = '0'
        _entryError = ''
        while choice =='0':
            logger.debug('channel_table() choice is ' + choice)
            DEVICE.open()

            # Does channel table file exist? 

            new_channel_table_fullpath = 'C:\R_S\instr\catv\channel_tables\\'\
                + new_channel_table + '.CHT'
            logger.debug('Query for file: \'' + new_channel_table_fullpath + '\'')
            
            try:
                query = "MMEM:CAT? \'"+new_channel_table_fullpath+"\'"
                logger.debug('Query: ' + query)
                dir_list = DEVICE.query(query)
                DEVICE.close()
                logger.debug(dir_list)
                if (dir_list == '\''+new_channel_table + '.CHT\'') :
                    choice = '4'
                    _channel_table = new_channel_table
                    logger.debug\
                        ('ETL reported that Channel Table \'' + \
                        new_channel_table + '\' exists')
                    section = 'UserEntered'
                    keyword = '_channel_table'
                    keyvalue = _channel_table
                    update_config_file(INI_FILENAME, section, keyword, keyvalue)
                    logger.debug('Wrote '+ _channel_table+' to INI file')
            except:                
                DEVICE.close()
                choice ='0'
                inputPrompt = \
                    """  ETL reported that Channel Table \'"""\
                        + new_channel_table + """\' does NOT exist.
                
                1 — continue with Channel Table that does not exist
                2 — enter a new Channel Table name
                3 — use default \'""" + _default_channel_table + \
                    """\'
                Q — Quit
                """ + _entryError + """
                Type a choice and press ENTER or ENTER alone to try again"""
                logger.debug(inputPrompt)
                choice = input (inputPrompt + ' >> ')
                _entryError = ''

                if choice == "q":
                    logger.debug('User chose to quit.')
                    done()
                elif choice == "Q":
                    logger.debug('User chose to quit.')
                    done()
                elif choice == "1":
                    logger.warning('User chose 1 to continue with \
Channel Table that does not exist.')
                elif choice == "2":
                    
                    inputPrompt = 'New Channel Table name'
                    new_channel_table = input (inputPrompt + ' >> ')
                    logger.debug\
                        ('User chose 2 to enter new Channel Table: \''+\
                            new_channel_table + '\'' )
                    # Delete characters that are not legal in file names
                    new_channel_table = re.sub(r'[\\/*?:"<>|]',"",\
                        new_channel_table)
                    choice = '0'
                elif choice == "3":
                    new_channel_table = _default_channel_table
                    logger.debug('User chose 3 to use default \'' + \
                        new_channel_table + '\'.')
                    choice = '0'
                elif choice == "":
                    logger.debug('User chose to try again ')
                    choice = '0'
                else:
                    _entryError = 'Your entry "'+choice+\
                        '" is not a valid choice'
                    choice = '0' 
                pass  # Keep going
    return
# ———————————————————————————————————————————————————
#              MENU
# ———————————————————————————————————————————————————
def menu():
    
    global _entryError
    global timedatewarning
    global _file
    global _ext
    global withoutConnection
    
    GpsMsg = '** No connection to ' + ipaddr.replace('\n','') + ' **'
 


    choice ='0'
    while choice =='0':    
        if withoutConnection == 0 :
            gps()
            if _gps == 1 :
                GpsMsg = ('''       GPS  
                          Lat, Lon: ''' + _lat.replace("\n","") + ''', '''\
                               + _lon.replace("\n","") + '''
                          Altitude: ''' + _altitude.replace("\n","") +\
                              ''' m
                          Number of satellites: ''' + \
                              _gpsValid.replace("\n","") ) 
            else:
                GpsMsg = "** GPS NOT CONNECTED **"
            
        _AzToTx = calcAzToTx()
            
        inputPrompt = ("""**** MAIN MENU ****\
    Control Script for Rohde & Schwarz ETL Analyzer"""
 + timedatewarning + """
              2 — enter Call Sign ["""+_callsign+"""] and Channel ["""+\
                  _channel+"""]
              3 — enter Measuring Location ["""+_MEASPT+"""]
              4 — enter Test position ["""+_testposition+"""]
              5 — enter Truck Heading ["""+str(_truckheading)+"""]
                    Truck antenna to transmitter ["""+str(_AzToTx)+"""]
              6 — Overview
              7 — Spectrum
              8 — Constellation Diagram
              9 — Modulation Errors
             10 — Eye Diagram
             11 — Echo Pattern
             12 — Capture Screen Shots
             13 — Measure Log — START CAPTURE
             14 — Site Checklist
             15 — Screen Shot of current screen
             
              Q — Quit
              
             """ + GpsMsg + """
              
              """ + _entryError + """
             Type a choice and press ENTER""") 
          

          
        choice = input (inputPrompt + ' >> ')
          
        _entryError = ''

        if choice == "q":
            done()
        elif choice == "Q":
            done()
        elif choice == "15":
            if withoutConnection == 0 :
                _file = 'SetByUser'
                _ext = 'PNG'
                PrintToFile()
            choice = '0' 
        elif choice == "14":
            ChecklistMenu()
            choice = '0' 
        elif choice == "13":
            if withoutConnection == 0 :
                MeasureLog()
            choice = '0' 
        elif choice == "12":
            if withoutConnection == 0 :
                CaptureScreenShots()
            choice = '0' 
        elif choice == "11":
            if withoutConnection == 0 :
                EchoPattern()
            choice = '0' 
        elif choice == "10":
            if withoutConnection == 0 :
                EyeDiagram()
            choice = '0' 
        elif choice == "9":
            if withoutConnection == 0 :
                ModulationErrors()
            choice = '0' 
        elif choice == "8":
            if withoutConnection == 0 :
                ConstDiagram()
            choice = '0' 
        elif choice == "7":
            if withoutConnection == 0 :
                spectrum()
            choice = '0' 
        elif choice == "6":
            if withoutConnection == 0 :
                overview()
            choice = '0' 
        elif choice == "5":
            requestTruckHeading()
            choice = '0' 
        elif choice == "4":
            requesttest()
            choice = '0' 
        elif choice == "3":
            requestmonpoint()
            choice = '0' 
        elif choice == "2":
            requestCallChan()
            choice = '0' 
        elif choice == "1":
            if withoutConnection == 0 :
                SetETLClock()
            choice = '0' 
        elif choice == 'n' :  # Not listed in menu
            checkcomputertime()
            choice = '0'
        elif choice == 'N' :  # Not listed in menu
            checkcomputertime()
            choice = '0'
        else:
            _entryError = 'Your entry "'+choice+'" is not a valid choice'
            choice = '0' 
            
    return
              


def requestCallChan():
    global _callsign
    _input = input("Call sign? ("+_callsign+") >> ")
    if _input != "" :
        # Replace space, convert to upper case
        _callsign =  _input.replace(" ","").upper()
        # Eliminate illegal characters in file names, limit to 9 char:
        _callsign = re.sub(r'[\\/*?:"<>|]',"",_callsign[:9])
        section , keyword , keyvalue = 'UserEntered' , '_callsign' ,\
             _callsign
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
        logger.debug('Call Sign is '+ _callsign)

    global _channel
    
    message = "Channel number? ("+_channel+") >> "
    chanEntered = 0
    while chanEntered == 0 :
        _input = input(message)
        chanEntered = 1
        message = "Channel number? ("+_channel+") >> "
        if _input == "" :
            return
        else:
            try:
                _input = int(_input)
            except ValueError:
                message = message + """
                MUST BE AN INTEGER >> """
                chanEntered = 0
                
                continue
               
            else:
                if  _input > 36 :
                    message = message + """
                    MUST BE LESS THAN 37 >> """
                    chanEntered = 0
                elif _input < 2 :
                    message = message + """
                    MUST BE GREATER THAN 1 >> """
                    chanEntered = 0
                else:
                    _channel = str(_input)
                    chanEntered = 1
                    section , keyword , keyvalue = 'UserEntered' , '_channel' \
                        , _channel
                    update_config_file(INI_FILENAME, section, keyword, keyvalue)
                    return _channel
                    #break
    return

def requestradial() :
    
    global _radial
    
    message = "Radial? (" + str(_radial)+ ") >> "
    while True:
        try:
            _input = int(input(message))
            message = "Radial? (" + str(_radial)+ ") >> "
        except ValueError:
            message = message + """
            MUST BE AN INTEGER >> """
            
            continue
           
        else:
            if  _input > 360 :
                message = message + """
                MUST BE LESS THAN 360 >> """
            elif _input < 0 :
                message = message + """
                MUST BE GREATER THAN 0 >> """
            else:
                _radial = _input
                
                section , keyword , keyvalue = 'UserEntered' , '_radial' , \
                    str(_radial)
                update_config_file(INI_FILENAME, section, keyword, keyvalue)
                return _radial
                break
            

def requestdist() :
    
    global _dist
    message = "Distance? (" + str(_dist)+ ") >> "
    while True:
        try:
            _input = int(input(message))
            message = "Distance? (" + str(_dist)+ ")"
        except ValueError:
            message = message + """
            MUST BE AN INTEGER >> """
            
            continue
           
        else:
            if  _input > 200 :
                message = message + """
                MUST BE LESS THAN 200 >> """
            elif _input < 0 :
                message = message + """
                MUST BE GREATER THAN 0 >> """
            else:
                _dist = _input
                section , keyword , keyvalue = 'UserEntered' , '_dist' , \
                    str(_dist)
                update_config_file(INI_FILENAME, section, keyword, keyvalue)
                return _dist
                break

def requestmonpoint():
    global _MEASPT
    requestradial()
    requestdist()
   
    _MEASPT = 'R' + str(_radial) + 'M' + str(_dist)
    logger.debug('Measurement point is ' + _MEASPT)
    return
     
def requesttest():
    global _testposition
    _input = input("Test? ("+_testposition+") >> ")
    if _input != "" :
        _testposition = _input.replace(" ","").upper()
        # Eliminate illegal characters in file names, limit length to first
        #   7 char:
        _testposition = re.sub(r'[\\/*?:"<>|]',"",_testposition[:7])        
        section , keyword , keyvalue = 'UserEntered' , '_testposition' , \
            _testposition
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
    logger.debug('Test is ' + _testposition)
    return

# ———————————————————————————————————————————————————
#              COMMON SETTINGS FOR ALL MEASUREMENTS
# ———————————————————————————————————————————————————

def commonsettings() :

    if (ping(ipaddr)) :    
        logger.debug("Common settings for all measurements")
        DEVICE.open()
        DEVICE.clear()
    
        # select the measurement mode "TV/Radio Analyzer/Receiver"
        DEVICE.write('INST CATV',)
        
        # Operation complete query. Returns an ASCII "+1"
        # when all pending overlapped operations have been completed
        _null = '0'
        try:
            _null = DEVICE.query('*OPC?',)
        except:
            logger.info("Select TV/Radio Analyzer mode — \
Operation Complete exception: " + _null.replace('\n','') )
            pass

        # Sets the TV standard
        #DEVICE.write('SET:TV:SpTAN ATSC')
        # Sets the constellation
        #DEVICE.write('SOURce:DM:FORM VSB')
        # Defines the symbol rate. 8VSB signal is by default 10.7622378 
        #   MSymbol/s.
        #DEVICE.write('DDEM:SRAT 10.7622378MA')
        # Selects and activates channel table; 
        try:
            DEVICE.write('CONF:TV:CTAB:SEL \"' + _channel_table + '\"',)
        except:
            logger.debug('Failed to select channel table \'' + \
                _channel_table +'\"',)
            pass
        #Specifies the pilot carrier frequency as reference frequency.
        #DEVICE.write('FREQ:PIL:STAT ON')
        # Tune to channel:
        DEVICE.write('FREQ:CHAN '+ _channel,)

        
        # Operation complete query. Returns an ASCII "+1"
        # when all pending overlapped operations have been completed
        _null = '0'
        try:
            _null = DEVICE.query('*OPC?',)
        except:
            logger.info("Tune to channel — Operation Complete failed: " \
                + _null.replace('\n','') )
            time.sleep(2)
            pass
        
        # Following step (RF Attenuator) produced screwy results
        # if we didn't wait for operation complete. 
        
        # Configures the RF attenuation mode to continuous
        # control of the attenuator.:
        logger.debug\
            ('Write to DEVICE: "SENS:POW:ACH:PRES:RLEV:AUTO SLA"')
        try:
            DEVICE.write('SENS:POW:ACH:PRES:RLEV:AUTO SLA',)
        except:
            logger.info("Set RF attenuation mode failed")
            pass
        try:
            _null = '0'
            _null = DEVICE.query('*OPC?',)
        except:
            logger.info\
                ("Set RF attenuation mode — Operation Complete failed: "\
                     + _null.replace('\n','') )
            time.sleep(2)
            pass
        
        global _frequency
        try:
            _frequency = DEVICE.query('FREQ:CENT?',) # in Hz
            logger.debug("Frequency reported by ETL: " + \
                _frequency.replace('\n','') + " Hz" )
        except:
            logger.info("Frequency query failed" )
            time.sleep(2)
            pass
        
        
        # Operation complete query. Returns an ASCII "+1" 
        # when all pending overlapped operations have been completed
        try:
            _null = DEVICE.query('*OPC?',) 
        except:
            logger.debug('Operation Complete query timeout')
            pass   

    else:
        logger.warning("No response from " + ipaddr.replace("\n","") )
    
    DEVICE.close()


    return

# ———————————————————————————————————————————————————
#              OVERVIEW
# ———————————————————————————————————————————————————


def overview():
    
    global _file , _ext
    _file = 'Overview'
    _ext = 'PNG'
    
    commonsettings()
    DEVICE.open()

    # Define the Digital TV Overview measurement type:
    DEVICE.write('CONF:DTV:MEAS OVER',)
    logger.debug('Requested Overview')
   
    # currently connected to the GPS receiver?
    try:
        _gps = int(DEVICE.query('SYST:POS:GPS:CONN?',))
        if _gps == 1 :    
            # Displays GPS data in the Overview measurement.
            DEVICE.write('DISP:MEAS:OVER:GPS:STAT ON',)
        else:
            logger.warning(' ————— GPS is NOT CONNECTED! —————')
    except :
        _entryError = \
            'Did not get GPS status. \
ETL may not be in "TV/Radio Analyzer" mode.'
        logger.debug(_entryError)
        pass


    # Operation complete query. Returns an ASCII "+1" 
    # when all pending overlapped operations have been completed
    try:
        _null = DEVICE.query('*OPC?',) 
    except:
        logger.debug('Operation Complete query timeout')
        pass   
    
    DEVICE.close()
    
    return

# ———————————————————————————————————————————————————
#              SPECTRUM
# ———————————————————————————————————————————————————

def spectrum() :
    
    global _file , _ext

    _file = 'Spectrum'
    _ext = 'PNG'

    commonsettings()
    
    DEVICE.open()

    # Define the Digital TV Spectrum measurement type:
    
    DEVICE.write('CONF:DTV:MEAS DSP',)
    logger.debug('Requested Spectrum')  
    # Operation complete query. Returns an ASCII "+1"
    # when all pending overlapped operations have been completed
    try:
        _null = DEVICE.query('*OPC?',)
    except:
        logger.debug\
            ('Operation Complete Query for setting Spectrum failed')
        DEVICE.close()
        time.sleep(1)
        DEVICE.open()
        pass
    # span  
    #DEVICE.write('FREQ:SPAN 10MHz')   
    # Operation complete query. Returns an ASCII "+1"
    # when all pending overlapped operations have been completed
    #_null = DEVICE.query('*OPC?',) 
    
    # Activates shoulder attenuation measurement in Spectrum measurement.
    DEVICE.write('CONF:DTV:MEAS:SATT ON',)
    logger.debug('Requested shoulder attenuation measurement')  
    #DEVICE.write('CONF:DTV:MEAS:SATT *RST')
    try:
        _null = DEVICE.query('*OPC?',)
    except:
        logger.debug\
            ('Operation Complete Query for setting shoulder attenuation \
measurement failed')
        pass
    
    # Returns the number of defined ranges.
    #NoOfRanges = DEVICE.query('ESP:RANG:COUNt?')
    #print('Number of Ranges '+NoOfRanges)
    
    # Returns the current position (number) of the reference range.
    #Range = DEVICE.query('ESP:RRAN?')
    #print('Current position of reference range '+Range)
    
    #Sets the RBW value to 10 kHz for the specified range.
    #DEVICE.write('SENS:ESP:RANG:BAND:RES 10000')
    
    # Sets the VBW for range 1 to 100 kHz
    #DEVICE.write('SENS:ESP:RANG:BAND:VID 100000')
    
    # Sets the sweep time for range 1 to 100 ms.
    #DEVICE.write('SENS:ESP:RANG:SWE:TIME 0.1')
    
    DEVICE.close()
    
    return

# ———————————————————————————————————————————————————
#              CONSTELLATION DIAGRAM
# ———————————————————————————————————————————————————

def ConstDiagram() :
    
    global _file , _ext
    _file = 'Constellation'
    _ext = 'PNG'
    
    commonsettings()
    DEVICE.open()

    # Define the Digital TV measurement type:
    DEVICE.write('CONF:DTV:MEAS CONS',)
    logger.debug('Requested Constellation Diagram')
    
    # Operation complete query. Returns an ASCII "+1" when all pending
    # overlapped operations have been completed
    try:
        _null = DEVICE.query('*OPC?',)
    except:
        logger.debug('Operation Complete Query for Constellation failed')
        pass
 

    DEVICE.close()
    
    return

# ———————————————————————————————————————————————————
#              MODULATION ERRORS
# ———————————————————————————————————————————————————

def ModulationErrors() :
    
    global _file
    global _ext
    _file = 'MER'
    _ext = 'PNG'
    
    commonsettings()
    DEVICE.open()

    # Define the Digital TV Modulation Errors measurement type:
    DEVICE.write('CONF:DTV:MEAS MERR',)
    logger.debug('Requested Modulation Errors')
    
    # Operation complete query. Returns an ASCII "+1" when all pending 
    # overlapped operations have been completed
    try:
        _null = DEVICE.query('*OPC?',)
    except:
        logger.debug\
            ('Operation Complete Query for Modulation Errors failed')
        pass


    DEVICE.close()
    
    return
  
# ———————————————————————————————————————————————————
#              EYE DIAGRAM
# ———————————————————————————————————————————————————
  
def EyeDiagram() :
    
    global _file
    global _ext
    _file = 'Eye'
    _ext = 'PNG'
    
    commonsettings()
    DEVICE.open()

    # Define the Digital TV measurement type:
    DEVICE.write('CONF:DTV:MEAS EYED',)
    logger.debug('Requested Eye Diagram')
    
    # Operation complete query. Returns an ASCII "+1" when all pending 
    # overlapped operations have been completed
    try:
        _null = DEVICE.query('*OPC?',) 
    except:
        logger.debug('Operation Complete query timeout')
        pass    

    DEVICE.close()
    
    return
    
# ———————————————————————————————————————————————————
#              ECHO PATTERN
# ———————————————————————————————————————————————————

def EchoPattern() :
    
    global _file
    global _ext
    _file = 'Echo'
    _ext = 'PNG'
    
    commonsettings()
    DEVICE.open()

    # Define the Digital TV measurement type:
    DEVICE.write('CONF:DTV:MEAS EPATtern',)
    logger.debug('Requested Echo Pattern')
    
    # Operation complete query. Returns an ASCII "+1" when all
    # pending overlapped operations have been completed
    try:
        _null = DEVICE.query('*OPC?',)
    except:
        logger.debug('Operation Complete Query for Echo Pattern failed')
        pass
 

    DEVICE.close()
    
    return

# ———————————————————————————————————————————————————
#              WAIT FOR DEMOD TO LOCK
# ———————————————————————————————————————————————————

def waitforlock() :
    logger.debug('Wait for demod to lock')
    DEVICE.open()
    _locked = 0
    _lockedcount = 0
    while _locked == 0 :
    # Queries if TV signal is available and locked.
    #  0 = not synchronized
    #  1 = synchronized
        try:
            _locked = int( DEVICE.query('CALC:DTV:RES:DEM:SYNC?',) )
        except:
            logger.debug('Locked query failed')
            pass

        #logger.debug('Locked status ' + str(_locked) )
        # Operation complete query. Returns an ASCII "+1" when all
        # pending overlapped operations have been completed        
        try:
            _null = DEVICE.query('*OPC?',)
        except:
            logger.debug\
                ('Operation Complete Query for Locked Query failed')
            pass 

        _lockedcount += 1  # Increment by 1 after each query
        #logger.debug("Locked count ="+ str(_lockedcount))
        global lockTimeOut
        lockTimeOut = 0  # Did not time out
        if _lockedcount >= 300 : 
            #Fail to lock after 300 queries (approx 3 sec.)
            lockTimeOut = 1  # Time out
            DEVICE.close()
            return      
    DEVICE.close()
    logger.debug("Locked count ="+ str(_lockedcount) +\
        " (approx 10 msec/count)")       
# ———————————————————————————————————————————————————
#              PRINT TO FILE
# ———————————————————————————————————————————————————
    
def PrintToFile() :
    global _file
    global _ext
    global _meas_results_folder

    logger.debug('Print ' + _file + ' to file')
# Typical file name:
# Z:\Measurement_results\WXXX\20200413_WXXX_R270M20\A\
#   20200413_WXXX_R270M20_A_Constellation.PNG

#Response = DEVICE.query('MMEM:CAT? "Z:\*"')
#print('CAT response ' + Response)

    # Get today's date as YYYYMMDD:
    _YYYYMMDD = datetime.datetime.now().strftime('%Y%m%d')

    # Time as HHMMSS:
    _HHMMSS = datetime.datetime.now().strftime('%H%M%S')
    
    _FILEPATH = _meas_results_folder + "\\"+_callsign+"\\"+_YYYYMMDD+"_"+\
        _callsign+"_"+_MEASPT+"\\"+_testposition
    _FILENAME = _callsign+"_"+_MEASPT+"_"+_testposition+"_"+_YYYYMMDD+"_"+\
        _HHMMSS+"_"+_file+"."+_ext
    

    if (ping(ipaddr)) :
        DEVICE.open()
        # Create directories
        DEVICE.write("MMEM:MDIR \'" + _meas_results_folder + "\\"+\
            _callsign+"\'",)
        DEVICE.write("MMEM:MDIR \'" + _meas_results_folder + "\\"+\
            _callsign+"\\"+_YYYYMMDD+"_"+_callsign+"_"+_MEASPT+"\'",)
        DEVICE.write("MMEM:MDIR \'"+_FILEPATH+"\'",)

        logger.debug(_FILEPATH+"\\"+_FILENAME)
        
            
        # Sets the colors for a printout or hardcopy
        #  1 — current screen colors with the background
        #      in white and the grid in black
        #  2 — optimized color set
        #  3 — user defined color set
        #  4 — current screen colors without any changes
        DEVICE.write('HCOP:CMAP:DEF4',)

        # Set the data format of the printout to PNG.
        DEVICE.write('HCOP:DEV:LANG '+_ext,)
        
        # color (not monochrome) hardcopy of the screen.
        DEVICE.write('HCOP:DEV:COL ON',)
        
        # Directs the hardcopy to a file.
        DEVICE.write("HCOP:DEST 'MMEM'",)
        
        #Selects the file name.
        DEVICE.write("MMEM:NAME \'"+_FILEPATH+"\\"+_FILENAME+"\'",)
        
        # comment text for the printout.
        DEVICE.write("HCOP:ITEM:WIND:TEXT \'" + _callsign + \
            "   Measuring Location: " + _MEASPT + "   Test: " +\
                 _testposition + "\'",)
        
        # Operation complete query. Returns an ASCII "+1" when all
        # pending overlapped operations have been completed
        try:
            _null = DEVICE.query('*OPC?',)
        except:
            logger.debug\
                ('Operation Complete Query for creating directories failed')
            pass
        
        # Saves the hardcopy output into the file
        try:
            logger.debug('Exporting ' + _file +' to ' + _FILENAME )
            DEVICE.write('HCOP',)
        except:
            logger.warning('Save to file failed')
        try:
            _null = DEVICE.query('*OPC?',)
        except:
            logger.debug\
                ('Operation Complete Query for saving hardcopy to file failed')
            pass
        try:
            dir_list = DEVICE.query("MMEM:CAT? \'"+_FILEPATH+"\\"+_FILENAME+\
                "\'")
            logger.warning('Exported ' + _file +' to ' + \
                dir_list.replace('\n',''))
        except:
            logger.debug("Failed to save \'"+_FILEPATH+"\\"+_FILENAME+"\'")
            logger.warning('!!! Failed to save ' + _file +' to ' + _FILENAME\
                 + ' !!!')
            pass
    else:
        logger.warning('No response from ' + ipaddr.replace('\n','') )
    DEVICE.close()    
    
    _file = 'SetByUser'
    _ext = 'PNG'
    return
    
# ———————————————————————————————————————————————————
#              CAPTURE SCREEN SHOTS
# ———————————————————————————————————————————————————


def CaptureScreenShots() :
    
    
    overview()
    waitforlock()
    # Wait for 5 seconds
    time.sleep(5)
    PrintToFile()  # Even if there is no lock, capture
    
    spectrum()
    # Wait for 5 seconds
    time.sleep(5)
    PrintToFile()
    
    ConstDiagram()
    waitforlock()
    if lockTimeOut == 0 : # 0 means demod is locked
        # Wait for 5 seconds
        time.sleep(5)
        PrintToFile()
    else:
        logger.warning("No Lock. Did NOT capture "+_file+".")
    
    ModulationErrors()
    waitforlock()
    if lockTimeOut == 0 : # 0 means demod is locked
        # Wait for 5 seconds
        time.sleep(5)
        PrintToFile()
    else:
        logger.warning("No Lock. Did NOT capture "+_file+".")
    
    EyeDiagram()
    waitforlock()
    if lockTimeOut == 0 : # 0 means demod is locked
        # Wait for 5 seconds
        time.sleep(5)
        PrintToFile()
    else:
        logger.warning("No Lock. Did NOT capture "+_file+".")
    
    EchoPattern()
    waitforlock()
    if lockTimeOut == 0 : # 0 means demod is locked
        # Wait for 5 seconds
        time.sleep(5)
        PrintToFile()
    else:
        logger.warning("No Lock. Did NOT capture "+_file+".")
    
    return 

# ———————————————————————————————————————————————————
#              MEASURE LOG
# ———————————————————————————————————————————————————

def MeasureLog() :
    
    logger.debug('Begin MeasureLog')
    global tdelta
    try:
        DEVICE.open()
        try:
            ETLdate = ''
            ETLdate = DEVICE.query('SYST:DATE?',) # 2020,5,9
        except:
            logger.debug('Failed to get ETL date')
            pass
        try:
            ETLtime = ''
            ETLtime = DEVICE.query('SYST:TIME?',) # 7,38,53
        except:
            logger.debug('Failed to get ETL time')
            pass
    
        try:
            ETLdatetime = str(datetime.datetime.strptime(ETLdate + ' ' + \
                ETLtime,'%Y,%m,%d\n %H,%M,%S\n'))
            logger.debug('ETL date time is '+ ETLdatetime)
            tdelta = datetime.datetime.strptime(ETLdatetime,'%Y-%m-%d %X').\
                timestamp() - datetime.datetime.now().timestamp()

            if (abs(tdelta) > 2) :
                logger.debug('ETL clock is ' + str(tdelta) + \
                    ' seconds different than computer')
                DEVICE.close()
                SetETLClock()
                DEVICE.open()
        except:
            logger.debug('Failed to compare ETL time to computer')
            pass


        # currently connected to the GPS receiver?
        try:
            _gps = int(DEVICE.query('SYST:POS:GPS:CONN?',) )
        except:
            logger.debug('Query for GPS status failed')
            pass
        logger.debug('GPS connected status is ' + str(_gps) )
        if _gps == 1 :    
            # Displays GPS data in the Overview measurement.
            DEVICE.write('DISP:MEAS:OVER:GPS:STAT ON',)
            # Activate measurement log. The overall performance of R&S
            # DEVICE is slightly affected if the measurement log is 
            # activated.
            DEVICE.write('CONF:MLOG ON',)
            
            # Operation complete query. Returns an ASCII "+1" when all
            # pending overlapped operations have been completed
            _null = '0'
            try:
                _null = DEVICE.query('*OPC?',)
            except:
                logger.debug\
                    ('Operation Complete Query for Measure Log failed')
                pass
            logger.debug('Measure log activated — operation complete? ' + \
                _null.replace('\n','') )
            DEVICE.close()  
            
            # Switch to Overview 
            # CSV file will be empty if Spectrum is on display of ETL.
            overview()  
            logger.info('Frequency reported by ETL: '+ \
                _frequency.replace('\n','') + ' Hz')
            # string with the syntax DD.MM.YYYY,HH:MM:SS
            # The start time within the measurement log from which the 
            # data is exported to a file.
            now = datetime.datetime.now()
            now_plus_3 = now + datetime.timedelta(seconds = 4) 
            # 3 seconds wasn't enough for GPS to settle
            StartTime = (now_plus_3).strftime('%d.%m.%Y,%H:%M:%S')

            logger.info('Capture start time will be ' + StartTime)
        
            inputPrompt = ("""Capture will start at """ + \
                (now_plus_3).strftime('%X') +
            """
            How many seconds?
            Strike ENTER key to end now.""")
            mTimer = 0
            choice = input(inputPrompt + ' >> ')
            logger.debug('User entered capture duration of \'' + choice +\
                 '\'')
            if choice == "":
                mTimer = 0
            else:
                            
                try:
                    mTimer = int(choice)
                except ValueError:  # choice is not an integer
                    mTimer = 0
            
            while mTimer > 0 :

                mins, secs = divmod(mTimer, 60)
                timeformat = '{:02d}:{:02d}'.format(mins, secs)
                print(datetime.datetime.now().strftime(FMT) + \
                    " Capture will end in "+ timeformat, end='\r', )

                time.sleep(1) # Wait for 1 secondw
                mTimer -= 1  # Decrement by 1 
        
            # string with the syntax DD.MM.YYYY,HH:MM:SS
            # The stop time within the measurement log until which the
            # data is exported to a file.
            StopTime = datetime.datetime.now().\
                strftime('%d.%m.%Y,%H:%M:%S')
            

            # Get today's date as YYYYMMDD:
            _YYYYMMDD = datetime.datetime.now().strftime('%Y%m%d')

            # Time as HHMMSS:
            _HHMMSS = datetime.datetime.now().strftime('%H%M%S')
            _file = 'capture'
            _ext = 'CSV'
            _FILEPATH = _meas_results_folder + "\\"+_callsign+"\\"\
                +_YYYYMMDD+"_"+_callsign+"_"+_MEASPT+"\\"+_testposition
            _FILENAME = _callsign+"_"+_MEASPT+"_"+_testposition+"_"\
                +_YYYYMMDD+"_"+_HHMMSS+"_"+_file+"."+_ext
            
            DEVICE.open()
            # Create directories
            DEVICE.write("MMEM:MDIR \'" + _meas_results_folder + "\\"\
                +_callsign+"\'",)
            DEVICE.write("MMEM:MDIR \'" + _meas_results_folder + "\\"\
                +_callsign+"\\"+_YYYYMMDD+"_"+_callsign+"_"+_MEASPT+"\'",)
            DEVICE.write("MMEM:MDIR \'"+_FILEPATH+"\'",)
            
            # export all available measurement values to a comma separated
            # values (.csv) file.
            # MMEMory:STORe:MLOG:DATA <StartTime>, <StopTime>, <ComprLev>,
            #   <FileName>
            # Compression level 0: one value is captured each second
            DEVICE.write('MMEM:STOR:MLOG:DATA \"'+StartTime+'\", \"'\
                +StopTime+'\", \"0\", \"'+_FILEPATH+'\\'+_FILENAME+'\"',)
            logger.info('Capture stop time is ' + StopTime )
            _null = '0'
            try:
                _null = DEVICE.query('*OPC?',)
            except:
                logger.debug('Operation Complete Query for Export to \
                    CSV failed')
                pass

            try:
                dir_list = DEVICE.query("MMEM:CAT? \'"+_FILEPATH+"\\"\
                    +_FILENAME+"\'")
                logger.warning('Exported ' + _file +' to ' + \
                    dir_list.replace('\n',''))
            except:
                logger.debug("Failed to save \'"+_FILEPATH+"\\"+_FILENAME\
                    +"\'")
                logger.warning('!!! Failed to save ' + _file +' to ' + \
                    _FILENAME + ' !!!')
                pass
            
            # Deactivate measurement log. The overall performance of R&S
            # ETL is slightly affected if the measurement log is activated.
            DEVICE.write('CONF:MLOG OFF',)
            DEVICE.close()

        else:
            logger.warning(' ————— GPS is NOT CONNECTED! —————')
            DEVICE.close()
    except :
        _entryError = 'Did not get GPS status. ETL may not be in \
"TV/Radio Analyzer" mode.'
        logger.debug(_entryError)
        DEVICE.close()
        pass

      
    return 


# ———————————————————————————————————————————————————
#              SET DEVICE CLOCK
# ———————————————————————————————————————————————————
def SetETLClock() :
    
    #global _entryError
    global timedatewarning

    
    if (ping(ipaddr)) :
        logger.warning('Setting ETL date and time.')
        
        # Sets the date for the internal calendar.
        # The sequence of entry is year, month, day.
        # Parameters:
        # DATE 1980 to 2099, 1 to 12, 1 to 31
        # Example: SYST:DATE 2000,6,1
        _DATE = datetime.datetime.now().strftime('%Y,%#m,%#d')
        DEVICE.open()
        DEVICE.write('SYST:DATE '+_DATE,)
        
        # Sets the internal clock. The sequence of entry is hour,
        # minute, second.
        # Parameters:
        # TIME Range: 0 to 23, 0 to 59, 0 to 59
        # Example: SYST:TIME 12,30,30
        _TIME = datetime.datetime.now().strftime('%#H,%#M,%#S')
        DEVICE.write('SYST:TIME '+_TIME,)
        
        try:
            ETLdate = DEVICE.query('SYST:DATE?',) # 2020,5,9
            ETLtime = DEVICE.query('SYST:TIME?',) # 7,38,53
        except:
            logger.debug('Query for ETL date and time failed')
            pass

        #ETLdatetime = str(datetime.datetime.strptime(ETLdate + ' ' + 
        #   ETLtime,'%Y,%m,%d %H,%M,%S')) 
        # This throws "raise ValueError("unconverted data remains: %s" %"
        ETLdatetime = str(datetime.datetime.strptime(ETLdate + ' ' + \
            ETLtime,'%Y,%m,%d\n %H,%M,%S\n')) 
        logger.info('ETL date and time: ' + ETLdatetime)
        tdelta = datetime.datetime.strptime(ETLdatetime,'%Y-%m-%d %X').\
            timestamp() - datetime.datetime.now().timestamp()
        timedatewarning = ''
        if (abs(tdelta) > 2) :
            timedatewarning = '  *** WARNING *** ' + '  Time delta is ' + \
                str(round(tdelta,1)) + ' seconds'
            logger.info('ETL date and time: ' + ETLdatetime + '   ' + \
                timedatewarning ,)

        # Operation complete query. Returns an ASCII "+1" when all
        # pending overlapped operations have been completed
        _null = DEVICE.query('*OPC?',)
        try:
            _null = DEVICE.query('*OPC?',)
        except:
            logger.debug\
                ('Operation Complete Query for Set ETL Clock failed')
            pass
        DEVICE.close()
    #else:
        #_entryError = 'FAILED to connect to DEVICE'

    return 

    
# ———————————————————————————————————————————————————
#              CHECKLIST MENU
# ———————————————————————————————————————————————————

def requestTruckHeading():
    logger.debug(' ******* TRUCK HEADING INPUT ******* ')
    global _truckheading
    global _AzToTx
    global _RevRadial
    
    message = "Truck Heading? ("+str(_truckheading)+") >> "
    while True:
        try:
            _input = int(input(message))
        except ValueError:
            message = message + """
                MUST BE AN INTEGER >> """
            
            continue
           
        else:
            logger.debug( 'Test if input is between 0 and 360' )
            if  _input > 360 :
                message = message + """
                MUST BE LESS THAN 360 >> """
            elif _input < 0 :
                message = message + """
                MUST BE GREATER THAN 0 >> """
            else:
                _truckheading = _input
                logger.debug( 'Truck heading typed as input was ' + \
                    str(_truckheading) )
                section = 'UserEntered'
                keyword = '_truckheading'
                keyvalue = str(_truckheading)
                update_config_file(INI_FILENAME, section, keyword, keyvalue)
                return
        
        

def calcAzToTx():
    
    global _AzToTx
    global _RevRadial
    global _radial
    global _truckheading
    try:
        _radial = int(_radial)
    except:
        pass
    else:
        if _radial >= 180 :
            _RevRadial = _radial - 180
        else: 
            _RevRadial = _radial + 180
        logger.debug('Reverse Radial is ' + str(_RevRadial) )

        
    try:
        _truckheading = int(_truckheading)
    except:
        pass
    else:
        try:
            _AzToTx = _RevRadial - _truckheading
        except:
            pass
        else:
            if ( _AzToTx < 0 ) :
                _AzToTx = 360 + ( _RevRadial - _truckheading )
            logger.debug( 'Truck antenna heading to transmitter is ' + \
                str(_AzToTx) )
    
    return _AzToTx #_truckheading
    

def requestTemperature():
    global _temperature
    _input = input("Temperature? ("+_temperature+") >> ")
    if _input != "" :
        _temperature =  _input
        section = 'UserEntered'
        keyword = '_temperature'
        keyvalue = _temperature
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
    return #ChecklistMenu()

def requestWind():
    global _wind 
    _input = input("Wind? ("+_wind+") >> ")
    if _input != "" :
        _wind =  _input
        section = 'UserEntered'
        keyword = '_wind'
        keyvalue = _wind
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
    return #ChecklistMenu()

def requestSkyCond():
    global _skycond 
    _input = input("Sky Conditions? ("+_skycond+") >> ")
    if _input != "" :
        _skycond =  _input
        section = 'UserEntered'
        keyword = '_skycond'
        keyvalue = _skycond
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
    return #ChecklistMenu()

def requestPrecip():
    global _precip
    _input = input("Precipitation? ("+_precip+") >> ")
    if _input != "" :
        _precip =  _input
        section = 'UserEntered'
        keyword = '_precip'
        keyvalue = _precip
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
    return #ChecklistMenu()

def requestTechName():
    global _techname
    _input = input("Technician name? ("+_techname+") >> ")
    if _input != "" :
        _techname =  _input
        section = 'UserEntered'
        keyword = '_techname'
        keyvalue = _techname
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
    return #ChecklistMenu()

def requestClutter():
    global _clutter
    _entryError = ''
    
          
    choice = "0" 
    while choice == "0":

    
        inputPrompt = ("""**** CLUTTER CATEGORY ****
        ("""+_clutter+""")
        1 — Open land
        2 — Agricultural
        3 — Rangeland
        4 — Water
        5 — Forest land
        6 — Wetland
        7 — Residential
        8 — Mixed Urban / Buildings
        9 — Commercial / Industrial
        10 — Snow and Ice
        
          """ + _entryError ) + """
        Type a choice and press ENTER"""

        choice = input (inputPrompt + ' >> ')
          
        _entryError = ''
        logger.debug('Clutter Category menu choice = ' + choice)

        if choice == "10":
            _clutter = '10 — Snow and Ice'
            return #ChecklistMenu()
        elif choice == "9":
            _clutter = '9 — Commercial / Industrial'
            return #ChecklistMenu()
        elif choice == "8":
            _clutter = '8 — Mixed Urban / Buildings'
            return #ChecklistMenu()
        elif choice == "7":
            _clutter = '7 — Residential'
            return #ChecklistMenu()
        elif choice == "6":
            _clutter = '6 — Wetland'
            return #ChecklistMenu()
        elif choice == "5":
            _clutter = '5 — Forest land'
            return #ChecklistMenu()
        elif choice == "4":
            _clutter = '4 — Water'
            return #ChecklistMenu()
        elif choice == "3":
            _clutter = '3 — Rangeland'
            return #ChecklistMenu()
        elif choice == "2":
            _clutter = '2 — Agricultural'
            return #ChecklistMenu()
        elif choice == "1":
            _clutter = '1 — Open land'
            return #ChecklistMenu()
        elif choice == "" :
            return #ChecklistMenu()
        else:
            _entryError = 'Your entry "'+choice+'" is not a valid choice'
            logger.debug(_entryError) 
            choice = "0"

   
    #return #ChecklistMenu()

def requestAntDirUp():
    global _antdirup
    _input = input("Actual antenna direction — Mast up? ("+_antdirup+\
        ") >> ")
    if _input != "" :
        _antdirup =  _input
        section = 'UserEntered'
        keyword = '_antdirup'
        keyvalue = _antdirup
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
    return #ChecklistMenu()

def returnAntDirDown():
    global _antdirdown
    _input = input("Actual antenna direction — Mast stowed? ("+\
        _antdirdown+") >> ")
    if _input != "" :
        _antdirdown =  _input
        section = 'UserEntered'
        keyword = '_antdirdown'
        keyvalue = _antdirdown
        update_config_file(INI_FILENAME, section, keyword, keyvalue)
    return #ChecklistMenu()

def ChecklistMenu() :
    
    _entryError = ''
    choice ='0'
    while choice =='0':
        inputPrompt = ("""**** SITE CHECKLIST ****
          1 — enter Temperature ["""+str(_temperature)+"""]
          2 — enter Wind ["""+_wind+"""]
          3 — enter Sky Conditions ["""+_skycond+"""]
          4 — enter Precip ["""+_precip+"""]
          5 — enter Technician Name ["""+_techname+"""]
          6 — enter Clutter Category ["""+_clutter+"""]
          7 — enter Actual Antenna Dir — Mast 30' ["""+\
              str(_antdirup)+"""]
          8 — enter Actual Antenna Dir — Mast stowed ["""+\
              str(_antdirdown)+"""]
          9 — Return to main menu
         10 — SAVE and EXPORT checklist to Excel file
         """ + _entryError +"""
         Type choice and press ENTER""")

    
        
        choice = input (inputPrompt + ' >> ')

        _entryError = ''
        
        if choice == "10":
            Checklist()
            return
        elif choice == "9":
            
            return
        elif choice == "":
            
            return
        elif choice == "8":
            returnAntDirDown()
            choice ='0'
        elif choice == "7":
            requestAntDirUp()
            choice ='0'
        elif choice == "6":
            requestClutter()
            section = 'UserEntered'
            keyword = '_clutter'
            keyvalue = _clutter
            update_config_file(INI_FILENAME, section, keyword, keyvalue)
            choice ='0'
        elif choice == "5":
            requestTechName()
            choice ='0'
        elif choice == "4":
            requestPrecip()
            choice ='0'
        elif choice == "3":
            requestSkyCond()
            choice ='0'
        elif choice == "2":
            requestWind()
            choice ='0'
        elif choice == "1":
            requestTemperature()
            choice ='0'
        else:
            _entryError = 'Your entry "'+choice+'" is not a valid choice'
            logger.debug('Site Checklist: ' + _entryError)
            choice ='0'
        #ChecklistMenu()
    return 
    

# ———————————————————————————————————————————————————
#              CHECKLIST (EXCEL SPREADSHEET)
# ———————————————————————————————————————————————————

def Checklist() :
    from openpyxl import Workbook 
    from openpyxl import load_workbook
    global _etlIDN
    if not withoutConnection == 0 :
        _etlIDN = 'Not connected'
    global _checklist_template

    workbook = Workbook()

    if os.path.isfile(_checklist_template.replace('\\','\\\\')) != 1 : 
        # If file doesn't exist
        logger.info( 'Checklist template file ' + _checklist_template + \
            " does not exist")
        # —————————————————
        choice ='0'
        while choice =='0':
                 
            inputPrompt = """Checklist template file """ + \
                _checklist_template + """ does not exist.
            Press ENTER to create a checklist without using a template or 
            type 1 then ENTER to specify a different template file."""
            choice = input (inputPrompt + ' >> ')
            
            if choice == '1':
                logger.debug('User chose to enter new template file.')
                inputPrompt = 'Enter a new template file'
                _checklist_template = input (inputPrompt + ' >> ')
                logger.debug('User entered new template file ' + \
                    _checklist_template)
                if os.path.isfile(_checklist_template.replace('\\','\\\\')) != 1 : # If file doesn't exist
                    logger.warning\
                        ( 'Newly entered checklist template file ' + \
                            _checklist_template + " does not exist")
                    choice = '0'

                else:  # File exists
                    section = 'ChecklistTemplate'
                    keyword = '_checklist_template'
                    keyvalue = _checklist_template
                    update_config_file\
                        (INI_FILENAME, section, keyword, keyvalue)
                    logger.debug\
                        ('_checklist_template written to INI file is '+ \
                            _checklist_template)
                    workbook = load_workbook\
                        (filename=_checklist_template.replace('\\','\\\\'))
            else:                    
                # Select first available sheet
                sheet = workbook.active
                sheet.title = 'Sheet1'
                
                sheet["A1"] = "Station:"
                sheet["A2"] = "TV Channel Number"
                sheet["A3"] = "Radial:"
                sheet["A4"] = "Measurement"
                sheet["A5"] = "Lat:"
                sheet["A6"] = "Lon:"
                sheet["A7"] = "Site Elev (m)"
                sheet["A8"] = "Heading to Tx (Calc):"
                sheet["A9"] = 'Frequency (MHz):'  # Added '(MHz)'
                sheet["B9"] = (float(_frequency)/1000000) 
                # Convert from Hz to MHz
                sheet["A10"] = "1/4 Wave"
                sheet["A11"] = "Date:"
                sheet["A12"] = "Time:"
                sheet["A13"] = "MEASUREMENTS"
                sheet["A14"] = "Enter Truck Heading:"
                sheet["A15"] = "Confirm Freq:"
                sheet["A16"] = "Confirm Ant. Length:"
                sheet["A17"] = "Location A"
                sheet["A18"] = "Location B"
                sheet["A19"] = "Temperature"
                sheet["A20"] = "Wind"
                sheet["A21"] = "Sky Conditions"
                sheet["A22"] = "Precip"
                sheet["A23"] = "Technician"
                sheet["C14"] = "Calculated Point AZ:"
                sheet["D16"] = "Actual Antenna Dir."
                sheet["C17"] = "Mast 30'"
                sheet["C18"] = "Mast Stowed"     
    else:
        workbook = load_workbook(filename=_checklist_template.replace\
            ('\\','\\\\'))
        
    # sheet
    # Select first available sheet
    sheet = workbook.active
    # sheet.title = 'Sheet1'
    
    sheet["B1"] = _callsign
    sheet["B2"] = int(_channel)
    sheet["B3"] = _radial
    sheet["B4"] = _dist
    sheet["B5"] = _lat
    sheet["B6"] = _lon
    sheet["B7"] = _altitude
    sheet["B8"] = _RevRadial
    sheet["B11"] = datetime.datetime.now().strftime('%m/%d/%Y')
    sheet["B12"] = datetime.datetime.now().strftime('%H:%M:%S')

    
    # Overwrite existing label in template:
    sheet["C19"] = 'Clutter category'
    sheet["C20"] = ''
    sheet["C21"] = ''
    
    sheet["B14"] = _truckheading
    sheet["B15"] = float(_frequency)/1000000  # Convert from Hz to MHz
    sheet["B19"] = _temperature
    sheet["B20"] = _wind
    sheet["B21"] = _skycond
    sheet["B22"] = _precip
    sheet["B23"] = _techname
    sheet["D14"] = _AzToTx
    sheet["D17"] = _antdirup
    sheet["D18"] = _antdirdown
    sheet["D19"] = _clutter
    sheet["A32"] = 'Test instrument:'
    sheet["B32"] = _etlIDN

    # Get today's date as YYYYMMDD:
    _YYYYMMDD = datetime.datetime.now().strftime('%Y%m%d')
    
    # Time as HHMMSS:
    _HHMMSS = datetime.datetime.now().strftime('%H%M%S')

    # Output file
    # C:\Shared\Measurement_results\XXXX\20200416_XXXX_RxxxMxx\Site Info\
    #   XXXX_RxxxMxx_Checklist.xlsx

    _FILEPATH = 'C:\\Shared\\Measurement_results\\'+_callsign+'\\'+\
        _YYYYMMDD+'_'+_callsign+'_'+_MEASPT+'\\Site Info'
    _FILENAME = _callsign+"_"+_MEASPT+"_"+_YYYYMMDD+"_"+_HHMMSS+\
        "_Checklist.xlsx"

    # Create a directory, first check if it already exists   
    if not os.path.exists('C:\\Shared\\Measurement_results\\'+_callsign):
        os.makedirs('C:\\Shared\\Measurement_results\\'+_callsign)
        
    if not os.path.exists('C:\\Shared\\Measurement_results\\'+_callsign+\
        '\\'+_YYYYMMDD+'_'+_callsign+'_'+_MEASPT):
        os.makedirs('C:\\Shared\\Measurement_results\\'+_callsign+'\\'+\
            _YYYYMMDD+'_'+_callsign+'_'+_MEASPT)

    if not os.path.exists(_FILEPATH):
        os.makedirs(_FILEPATH)
        
    if not os.path.exists(_FILEPATH + '\\Photos' ):
        os.makedirs(_FILEPATH + '\\Photos' )
    
    workbook.save( filename= _FILEPATH + "\\" + _FILENAME )
    if os.path.isfile(_FILEPATH + "\\" + _FILENAME) == 1 : # If file exists
        logger.warning('Exported checklist to ' + _FILENAME )
    else:
        logger.warning('FAILED to save ' + _FILENAME )
    
    return

# ———————————————————————————————————————————————————
#              DONE
# ———————————————————————————————————————————————————
def done() :
    try:
        _lat
    except NameError:
        pass
    else:
        logger.info('Latitude, Longitude: '+_lat.replace("\n","")+','+\
            _lon.replace("\n",""))
    logger.info('Quit')
    sys.exit(0) 


# ———————————————————————————————————————————————————
#              MAIN
# ———————————————————————————————————————————————————

def main() :
    
    start_logging()
    setvars()
    logger.debug('main() returned from setvars()')
    ip_ok()
    logger.debug('main() returned from ip_ok()')
    meas_results_folder()
    logger.debug('main() returned from meas_results_folder()')
    channel_table()
    logger.debug('main() returned from channel_table()')
    menu()

main()


